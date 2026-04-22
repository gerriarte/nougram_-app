"""
Onboarding Service - Business logic for onboarding operations
"""

import logging
from datetime import date
from decimal import Decimal
from io import BytesIO
from typing import Any

from sqlalchemy.ext.asyncio import AsyncSession
from sqlalchemy.orm.attributes import flag_modified

from app.core.calculations import calculate_blended_cost_rate
from app.core.currency import EXCHANGE_RATES_TO_USD
from app.core.sheets import get_sheets_client
from app.models.cost import CostFixed
from app.models.equipment import EquipmentAmortization
from app.models.team import TeamMember
from app.repositories.factory import RepositoryFactory
from app.repositories.organization_repository import OrganizationRepository
from app.schemas.onboarding import (
    CompleteOnboardingRequest,
    CountryCode,
    Currency,
    OnboardingExpense,
    OnboardingImportSheetsRequest,
    OnboardingTeamMember,
    TemporaryBCRRequest,
)

logger = logging.getLogger(__name__)


class OnboardingService:
    """Service for onboarding operations"""

    INVENTORY_TEMPLATE_CATALOG = [
        {
            "id": "laptop",
            "name": "Laptop de Trabajo",
            "amount": "800000",
            "currency": "COP",
            "category": "Tools",
            "amortizable": True,
            "icon": "💻",
            "preSelectedFor": ["dev", "design", "marketing"],
        },
        {
            "id": "monitor",
            "name": "Monitor Externo",
            "amount": "150000",
            "currency": "COP",
            "category": "Tools",
            "amortizable": True,
            "icon": "🖥️",
            "preSelectedFor": ["dev", "design"],
        },
        {
            "id": "laptop_high_end",
            "name": "Laptops High-End (MacBook Pro/Dell XPS)",
            "amount": "9500000",
            "currency": "COP",
            "category": "Tools",
            "amortizable": True,
            "icon": "💻",
            "preSelectedFor": ["dev", "design", "marketing", "consulting"],
        },
        {
            "id": "monitor_4k_profesional",
            "name": "Monitores 4K Profesionales",
            "amount": "2800000",
            "currency": "COP",
            "category": "Tools",
            "amortizable": True,
            "icon": "🖥️",
            "preSelectedFor": ["dev", "design", "marketing"],
        },
        {
            "id": "servidor_nas",
            "name": "Servidores NAS / Almacenamiento Local",
            "amount": "12000000",
            "currency": "COP",
            "category": "Tools",
            "amortizable": True,
            "icon": "🗄️",
            "preSelectedFor": ["dev", "consulting"],
        },
        {
            "id": "equipamiento_red_enterprise",
            "name": "Equipamiento de Red (Routers, Access Points Enterprise)",
            "amount": "6500000",
            "currency": "COP",
            "category": "Tools",
            "amortizable": True,
            "icon": "📡",
            "preSelectedFor": ["dev", "consulting"],
        },
        {
            "id": "equipos_videoconferencia",
            "name": "Equipos de Videoconferencia (Sala de juntas)",
            "amount": "5200000",
            "currency": "COP",
            "category": "Tools",
            "amortizable": True,
            "icon": "🎥",
            "preSelectedFor": ["consulting", "marketing", "dev", "design"],
        },
        {
            "id": "camaras_produccion",
            "name": "Camaras y Equipo de Produccion (Contenido Interno)",
            "amount": "8400000",
            "currency": "COP",
            "category": "Tools",
            "amortizable": True,
            "icon": "📷",
            "preSelectedFor": ["marketing", "design"],
        },
        {
            "id": "celulares_corporativos",
            "name": "Celulares",
            "amount": "3200000",
            "currency": "COP",
            "category": "Tools",
            "amortizable": True,
            "icon": "📱",
            "preSelectedFor": ["consulting", "marketing", "dev", "design"],
        },
        {
            "id": "camara_fotografica",
            "name": "Camara fotografica",
            "amount": "4500000",
            "currency": "COP",
            "category": "Tools",
            "amortizable": True,
            "icon": "📸",
            "preSelectedFor": ["marketing", "design"],
        },
        {
            "id": "flash_estudio",
            "name": "Flash de Estudio",
            "amount": "1700000",
            "currency": "COP",
            "category": "Tools",
            "amortizable": True,
            "icon": "⚡",
            "preSelectedFor": ["marketing", "design"],
        },
        {
            "id": "luces_led",
            "name": "Luces Led",
            "amount": "1200000",
            "currency": "COP",
            "category": "Tools",
            "amortizable": True,
            "icon": "💡",
            "preSelectedFor": ["marketing", "design"],
        },
        {
            "id": "computadora_escritorio",
            "name": "Computadora de Escritorio",
            "amount": "5600000",
            "currency": "COP",
            "category": "Tools",
            "amortizable": True,
            "icon": "🖥️",
            "preSelectedFor": ["dev", "design", "consulting"],
        },
        {
            "id": "microfonos",
            "name": "Microfonos",
            "amount": "900000",
            "currency": "COP",
            "category": "Tools",
            "amortizable": True,
            "icon": "🎙️",
            "preSelectedFor": ["marketing", "consulting"],
        },
        {
            "id": "escritorios_trabajo",
            "name": "Escritorios de trabajo",
            "amount": "1300000",
            "currency": "COP",
            "category": "Tools",
            "amortizable": True,
            "icon": "🪑",
            "preSelectedFor": ["dev", "design", "consulting", "marketing"],
        },
        {
            "id": "televisor",
            "name": "Televisor",
            "amount": "2600000",
            "currency": "COP",
            "category": "Tools",
            "amortizable": True,
            "icon": "📺",
            "preSelectedFor": ["consulting", "marketing"],
        },
        {
            "id": "video_camara",
            "name": "Video Camara",
            "amount": "5800000",
            "currency": "COP",
            "category": "Tools",
            "amortizable": True,
            "icon": "📹",
            "preSelectedFor": ["marketing", "design"],
        },
        {
            "id": "adobe_cc",
            "name": "Adobe Creative Cloud",
            "amount": "150000",
            "currency": "COP",
            "category": "Software",
            "amortizable": False,
            "icon": "🎨",
            "preSelectedFor": ["design", "marketing"],
        },
        {
            "id": "chatgpt_plus",
            "name": "ChatGPT Plus",
            "amount": "20",
            "currency": "USD",
            "category": "Software",
            "amortizable": False,
            "icon": "🤖",
            "preSelectedFor": ["dev", "marketing", "design"],
        },
        {
            "id": "figma",
            "name": "Figma Professional",
            "amount": "12",
            "currency": "USD",
            "category": "Software",
            "amortizable": False,
            "icon": "🎨",
            "preSelectedFor": ["design"],
        },
        {
            "id": "notion",
            "name": "Notion Pro",
            "amount": "8",
            "currency": "USD",
            "category": "Software",
            "amortizable": False,
            "icon": "📝",
            "preSelectedFor": ["consulting", "design"],
        },
        {
            "id": "google_workspace_enterprise",
            "name": "Google Workspace Enterprise",
            "amount": "27",
            "currency": "USD",
            "category": "Software",
            "amortizable": False,
            "icon": "📧",
            "preSelectedFor": ["consulting", "marketing", "dev", "design"],
        },
        {
            "id": "monday_com",
            "name": "Monday.com",
            "amount": "19",
            "currency": "USD",
            "category": "Software",
            "amortizable": False,
            "icon": "📊",
            "preSelectedFor": ["consulting", "marketing", "dev", "design"],
        },
        {
            "id": "asana_enterprise",
            "name": "Asana Enterprise",
            "amount": "31",
            "currency": "USD",
            "category": "Software",
            "amortizable": False,
            "icon": "✅",
            "preSelectedFor": ["consulting", "marketing", "dev", "design"],
        },
        {
            "id": "hubspot_marketing_sales_hub",
            "name": "HubSpot Marketing & Sales Hub",
            "amount": "200",
            "currency": "USD",
            "category": "Software",
            "amortizable": False,
            "icon": "🧲",
            "preSelectedFor": ["marketing", "consulting"],
        },
        {
            "id": "snowflake",
            "name": "Snowflake",
            "amount": "300",
            "currency": "USD",
            "category": "Software",
            "amortizable": False,
            "icon": "❄️",
            "preSelectedFor": ["dev", "consulting"],
        },
        {
            "id": "bigquery",
            "name": "BigQuery",
            "amount": "150",
            "currency": "USD",
            "category": "Software",
            "amortizable": False,
            "icon": "🗄️",
            "preSelectedFor": ["dev", "consulting"],
        },
        {
            "id": "tableau_enterprise",
            "name": "Tableau Enterprise",
            "amount": "75",
            "currency": "USD",
            "category": "Software",
            "amortizable": False,
            "icon": "📈",
            "preSelectedFor": ["consulting", "marketing"],
        },
        {
            "id": "semrush",
            "name": "Semrush",
            "amount": "140",
            "currency": "USD",
            "category": "Software",
            "amortizable": False,
            "icon": "🔎",
            "preSelectedFor": ["marketing"],
        },
        {
            "id": "ahrefs_agency",
            "name": "Ahrefs Agency",
            "amount": "249",
            "currency": "USD",
            "category": "Software",
            "amortizable": False,
            "icon": "🔗",
            "preSelectedFor": ["marketing"],
        },
        {
            "id": "slack_enterprise_grid",
            "name": "Slack Enterprise Grid",
            "amount": "25",
            "currency": "USD",
            "category": "Software",
            "amortizable": False,
            "icon": "💬",
            "preSelectedFor": ["consulting", "marketing", "dev", "design"],
        },
        {
            "id": "ai_apis_openai_anthropic",
            "name": "APIs de IA (OpenAI / Anthropic)",
            "amount": "200",
            "currency": "USD",
            "category": "Software",
            "amortizable": False,
            "icon": "🤖",
            "preSelectedFor": ["dev", "marketing", "consulting"],
        },
        {
            "id": "sprout_social",
            "name": "Sprout Social",
            "amount": "249",
            "currency": "USD",
            "category": "Software",
            "amortizable": False,
            "icon": "🌱",
            "preSelectedFor": ["marketing"],
        },
        {
            "id": "hootsuite_enterprise",
            "name": "Hootsuite Enterprise",
            "amount": "249",
            "currency": "USD",
            "category": "Software",
            "amortizable": False,
            "icon": "🦉",
            "preSelectedFor": ["marketing"],
        },
        {
            "id": "arriendo_oficina",
            "name": "Arriendo Oficina",
            "amount": "3500000",
            "currency": "COP",
            "category": "Overhead",
            "amortizable": False,
            "icon": "🏢",
            "preSelectedFor": ["consulting", "marketing", "dev", "design"],
        },
        {
            "id": "administracion_vigilancia",
            "name": "Administracion y Vigilancia",
            "amount": "600000",
            "currency": "COP",
            "category": "Overhead",
            "amortizable": False,
            "icon": "🛡️",
            "preSelectedFor": ["consulting", "marketing", "dev", "design"],
        },
        {
            "id": "servicios_publicos",
            "name": "Servicios Publicos (Luz, Agua, Gas)",
            "amount": "450000",
            "currency": "COP",
            "category": "Overhead",
            "amortizable": False,
            "icon": "💡",
            "preSelectedFor": ["consulting", "marketing", "dev", "design"],
        },
        {
            "id": "internet_fibra_dedicada",
            "name": "Internet Fibra Optica Dedicada",
            "amount": "350000",
            "currency": "COP",
            "category": "Overhead",
            "amortizable": False,
            "icon": "🌐",
            "preSelectedFor": ["consulting", "marketing", "dev", "design"],
        },
        {
            "id": "servicio_aseo_limpieza",
            "name": "Servicio de Aseo y Limpieza",
            "amount": "300000",
            "currency": "COP",
            "category": "Overhead",
            "amortizable": False,
            "icon": "🧹",
            "preSelectedFor": ["consulting", "marketing", "dev", "design"],
        },
        {
            "id": "seguros_multiriesgo_equipos",
            "name": "Seguros Multi-riesgo y Equipos",
            "amount": "280000",
            "currency": "COP",
            "category": "Overhead",
            "amortizable": False,
            "icon": "🧾",
            "preSelectedFor": ["consulting", "marketing", "dev", "design"],
        },
        {
            "id": "cloud_hosting_enterprise",
            "name": "Cloud Hosting (AWS / Azure / Vercel)",
            "amount": "900000",
            "currency": "COP",
            "category": "Overhead",
            "amortizable": False,
            "icon": "☁️",
            "preSelectedFor": ["dev", "consulting", "marketing"],
        },
        {
            "id": "outsourcing_contable_auditoria",
            "name": "Outsourcing Contable y Auditoria",
            "amount": "1200000",
            "currency": "COP",
            "category": "Overhead",
            "amortizable": False,
            "icon": "📚",
            "preSelectedFor": ["consulting", "marketing", "dev", "design"],
        },
        {
            "id": "retainer_legal_laboral",
            "name": "Retainer Legal y Laboral",
            "amount": "1500000",
            "currency": "COP",
            "category": "Overhead",
            "amortizable": False,
            "icon": "⚖️",
            "preSelectedFor": ["consulting", "marketing", "dev", "design"],
        },
        {
            "id": "reserva_recambio_hardware",
            "name": "Reserva Recambio de Hardware",
            "amount": "800000",
            "currency": "COP",
            "category": "Overhead",
            "amortizable": False,
            "icon": "🖥️",
            "preSelectedFor": ["dev", "design", "consulting"],
        },
        {
            "id": "coworking",
            "name": "Arriendo Coworking",
            "amount": "300000",
            "currency": "COP",
            "category": "Overhead",
            "amortizable": False,
            "icon": "🏢",
            "preSelectedFor": [],
        },
        {
            "id": "internet",
            "name": "Internet",
            "amount": "80000",
            "currency": "COP",
            "category": "Overhead",
            "amortizable": False,
            "icon": "🌐",
            "preSelectedFor": ["dev", "design", "marketing", "consulting"],
        },
        {
            "id": "hosting",
            "name": "Hosting Web",
            "amount": "50000",
            "currency": "COP",
            "category": "Overhead",
            "amortizable": False,
            "icon": "☁️",
            "preSelectedFor": ["dev"],
        },
    ]
    SUPPORTED_COUNTRY_CODES = {item.value for item in CountryCode}
    SUPPORTED_CURRENCY_CODES = {item.value for item in Currency}
    SUPPORTED_EXPENSE_CATEGORIES = {"rent", "software", "services"}

    def __init__(self, db: AsyncSession, organization_id: int):
        """
        Initialize OnboardingService

        Args:
            db: Database session
            organization_id: Organization ID for tenant scoping
        """
        self.db = db
        self.organization_id = organization_id
        self.org_repo = OrganizationRepository(db)
        self.team_repo = RepositoryFactory.create_team_repository(db, organization_id)
        self.cost_repo = RepositoryFactory.create_cost_repository(db, organization_id)

    async def get_inventory_templates(self) -> dict[str, Any]:
        """Return onboarding inventory template catalog from backend source."""
        return {
            "items": self.INVENTORY_TEMPLATE_CATALOG,
            "source": "backend_catalog",
        }

    async def get_onboarding_draft(self) -> dict[str, Any]:
        """Return onboarding draft from organization settings."""
        org = await self.org_repo.get_by_id(self.organization_id)
        if not org:
            raise ValueError(f"Organization {self.organization_id} not found")

        settings = org.settings or {}
        draft = settings.get("onboarding_draft") or {}
        if not isinstance(draft, dict):
            draft = {}
        return draft

    async def save_onboarding_draft(self, draft_data: dict[str, Any]) -> dict[str, Any]:
        """Persist onboarding draft in organization settings."""
        org = await self.org_repo.get_by_id(self.organization_id)
        if not org:
            raise ValueError(f"Organization {self.organization_id} not found")

        if org.settings is None:
            org.settings = {}

        org.settings["onboarding_draft"] = draft_data or {}
        flag_modified(org, "settings")
        await self.db.commit()
        await self.db.refresh(org)
        return org.settings.get("onboarding_draft") or {}

    async def get_benchmarks(
        self, profile_type: str, country: str = "US", currency: str = "USD"
    ) -> dict[str, Any]:
        """
        Get benchmark values for a business profile

        Args:
            profile_type: Profile type (freelance, company, agency)
            country: Country code
            currency: Currency code

        Returns:
            Dictionary with benchmark values
        """
        # Benchmarks hardcoded (en producción podrían venir de base de datos o API externa)
        benchmarks_data = {
            "freelance": {
                "avg_monthly_income": Decimal("5000"),
                "avg_margin": Decimal("25"),
                "avg_hours_per_month": Decimal("160"),
                "avg_team_size": None,
                "avg_salary": None,
                "avg_clients": None,
            },
            "company": {
                "avg_monthly_income": None,
                "avg_margin": Decimal("30"),
                "avg_hours_per_month": None,
                "avg_team_size": 5,
                "avg_salary": Decimal("3000"),
                "avg_clients": None,
            },
            "agency": {
                "avg_monthly_income": None,
                "avg_margin": Decimal("35"),
                "avg_hours_per_month": None,
                "avg_team_size": 10,
                "avg_salary": Decimal("3500"),
                "avg_clients": 5,
            },
        }

        # Ajustar según país (ejemplo: Colombia tiene salarios más bajos)
        if country == "COL":
            if benchmarks_data[profile_type].get("avg_salary"):
                benchmarks_data[profile_type]["avg_salary"] *= Decimal("0.4")  # Aproximado

        return {
            "profile_type": profile_type,
            "country": country,
            "currency": currency,
            "benchmarks": benchmarks_data.get(profile_type, {}),
            "source": "industry_standard",
        }

    async def complete_onboarding(self, request: CompleteOnboardingRequest) -> dict[str, Any]:
        """
        Complete onboarding by saving all configuration

        Args:
            request: CompleteOnboardingRequest with all onboarding data

        Returns:
            Dictionary with results of onboarding completion
        """
        try:
            request_currency = getattr(request.currency, "value", request.currency).upper()
            request_country = getattr(request.country, "value", request.country).upper()

            # Strict consistency: onboarding complete must persist one canonical currency.
            for member_data in request.team_members:
                member_currency = getattr(
                    member_data.currency, "value", member_data.currency
                ).upper()
                if member_currency != request_currency:
                    raise ValueError(
                        f"Mixed currencies are not allowed in onboarding team members. "
                        f"Expected {request_currency}, got {member_currency}."
                    )
            for expense_data in request.expenses:
                expense_currency = getattr(
                    expense_data.currency, "value", expense_data.currency
                ).upper()
                if expense_currency != request_currency:
                    raise ValueError(
                        f"Mixed currencies are not allowed in onboarding expenses. "
                        f"Expected {request_currency}, got {expense_currency}."
                    )
            for inventory_item in request.inventory_items or []:
                inventory_currency = str(inventory_item.get("currency") or "").upper()
                if inventory_currency and inventory_currency != request_currency:
                    raise ValueError(
                        f"Mixed currencies are not allowed in onboarding inventory items. "
                        f"Expected {request_currency}, got {inventory_currency}."
                    )

            # 1. Update organization
            org = await self.org_repo.get_by_id(self.organization_id)
            if not org:
                raise ValueError(f"Organization {self.organization_id} not found")

            # Update organization name if provided
            if request.organization_name:
                org.name = request.organization_name

            # Update primary currency
            if request_currency:
                if org.settings is None:
                    org.settings = {}
                org.settings["primary_currency"] = request_currency
                # Keep legacy key aligned to avoid mixed legacy/current settings.
                org.settings["currency"] = request_currency

            # Update settings with onboarding data
            if org.settings is None:
                org.settings = {}

            if request.organization_description:
                org.settings["description"] = request.organization_description

            org.settings["country"] = request_country
            org.settings["profile_type"] = request.profile_type
            org.settings["onboarding_completed"] = True
            org.settings["onboarding_draft"] = {}

            if request.tax_structure:
                org.settings["tax_structure"] = request.tax_structure
            if request.social_charges_config:
                org.settings["social_charges_config"] = request.social_charges_config
            if request.inventory_items:
                # Keep inventory detail for future modules (assets, amortization, audits).
                org.settings["onboarding_inventory_items"] = request.inventory_items

            # Mark settings as modified so SQLAlchemy detects the change
            flag_modified(org, "settings")

            await self.db.commit()
            await self.db.refresh(org)

            # 2. Create team members
            team_members_created = 0
            for member_data in request.team_members:
                # Convert monthly hours to weekly using the same 4.33 factor used in BCR engine.
                billable_hours_per_week = max(
                    1,
                    int(
                        round(Decimal(str(member_data.billable_hours_per_month)) / Decimal("4.33"))
                    ),
                )

                # Create TeamMember model instance
                team_member = TeamMember(
                    name=member_data.name,
                    role=member_data.role,
                    salary_monthly_brute=member_data.salary_monthly_brute,
                    currency=getattr(member_data.currency, "value", member_data.currency).upper(),
                    billable_hours_per_week=billable_hours_per_week,
                    apply_social_charges=True,
                    is_active=True,
                    organization_id=self.organization_id,
                )
                await self.team_repo.create(team_member)
                team_members_created += 1

            # 3. Create fixed costs (expenses)
            expenses_created = 0
            amortization_assets_created = 0
            for expense_data in request.expenses:
                expense_currency_code = getattr(
                    expense_data.currency, "value", expense_data.currency
                ).upper()
                expense_kind = getattr(expense_data, "cost_type", "operational")
                if expense_kind == "amortization":
                    # Financially strict: amortizable assets must be persisted as EquipmentAmortization.
                    category = "Software" if expense_data.category == "software" else "Hardware"
                    useful_life_months = 24 if category == "Software" else 36
                    exchange_rate_at_purchase = None
                    if expense_data.currency and expense_currency_code != "USD":
                        exchange_rate_at_purchase = Decimal(
                            str(EXCHANGE_RATES_TO_USD.get(expense_currency_code, 1.0))
                        )

                    asset = EquipmentAmortization(
                        name=expense_data.name,
                        description=f"Onboarding amortization asset: {expense_data.name}",
                        category=category,
                        purchase_price=expense_data.amount_monthly,
                        purchase_date=date.today(),
                        currency=expense_currency_code,
                        exchange_rate_at_purchase=exchange_rate_at_purchase,
                        useful_life_months=useful_life_months,
                        salvage_value=Decimal("0"),
                        depreciation_method="straight_line",
                        is_active=True,
                        organization_id=self.organization_id,
                    )
                    self.db.add(asset)
                    amortization_assets_created += 1
                    continue

                # Operational expense remains in CostFixed
                cost_fixed = CostFixed(
                    name=expense_data.name,
                    category=expense_data.category,
                    amount_monthly=expense_data.amount_monthly,
                    currency=expense_currency_code,
                    description=f"{expense_kind.capitalize()} expense: {expense_data.name}",
                    organization_id=self.organization_id,
                )
                await self.cost_repo.create(cost_fixed)
                expenses_created += 1

            # 3.1 Create amortization assets from onboarding inventory snapshot
            for inventory_item in request.inventory_items or []:
                try:
                    is_amortizable = bool(inventory_item.get("amortizable"))
                    if not is_amortizable:
                        continue

                    category_raw = str(inventory_item.get("category") or "")
                    category = "Software" if category_raw.lower() == "software" else "Hardware"
                    default_useful_life = 24 if category == "Software" else 36
                    useful_life_months = int(
                        inventory_item.get("useful_life_months") or default_useful_life
                    )
                    useful_life_months = max(1, useful_life_months)

                    quantity = int(inventory_item.get("quantity") or 1)
                    quantity = max(1, quantity)

                    purchase_price_raw = (
                        inventory_item.get("purchase_price")
                        if inventory_item.get("purchase_price") is not None
                        else inventory_item.get("amount_monthly")
                    )
                    purchase_price = Decimal(str(purchase_price_raw or "0")) * Decimal(
                        str(quantity)
                    )
                    if purchase_price <= 0:
                        continue

                    salvage_value = Decimal(str(inventory_item.get("salvage_value") or "0"))
                    if salvage_value < 0:
                        salvage_value = Decimal("0")
                    if salvage_value > purchase_price:
                        salvage_value = purchase_price

                    purchase_date_value = date.today()
                    purchase_date_raw = inventory_item.get("purchase_date")
                    if purchase_date_raw:
                        try:
                            purchase_date_value = date.fromisoformat(str(purchase_date_raw))
                        except Exception:
                            purchase_date_value = date.today()

                    depreciation_method = (
                        str(inventory_item.get("depreciation_method") or "straight_line")
                        .strip()
                        .lower()
                    )
                    if depreciation_method not in {"straight_line", "declining_balance"}:
                        depreciation_method = "straight_line"

                    item_currency = str(
                        inventory_item.get("currency") or request_currency or "USD"
                    ).upper()
                    exchange_rate_at_purchase = None
                    if item_currency != "USD":
                        exchange_rate_at_purchase = Decimal(
                            str(EXCHANGE_RATES_TO_USD.get(item_currency, 1.0))
                        )

                    asset = EquipmentAmortization(
                        name=str(inventory_item.get("name") or "Activo onboarding"),
                        description="Generated from onboarding inventory",
                        category=category,
                        purchase_price=purchase_price,
                        purchase_date=purchase_date_value,
                        currency=item_currency,
                        exchange_rate_at_purchase=exchange_rate_at_purchase,
                        useful_life_months=useful_life_months,
                        salvage_value=salvage_value,
                        depreciation_method=depreciation_method,
                        is_active=True,
                        organization_id=self.organization_id,
                    )
                    self.db.add(asset)
                    amortization_assets_created += 1
                except Exception:
                    # Keep onboarding resilient if one malformed inventory item is received.
                    continue

            await self.db.commit()

            # 4. Calculate BCR after saving
            bcr = await calculate_blended_cost_rate(
                self.db, primary_currency=request_currency, tenant_id=self.organization_id
            )

            logger.info(
                f"Onboarding completed for organization {self.organization_id}: "
                f"{team_members_created} team members, {expenses_created} operational expenses, "
                f"{amortization_assets_created} amortization assets"
            )

            return {
                "success": True,
                "message": "Onboarding completed successfully",
                "organization_id": self.organization_id,
                "team_members_created": team_members_created,
                "expenses_created": expenses_created,
                "bcr_calculated": str(bcr),
                "organization": {
                    "id": org.id,
                    "name": org.name,
                    "primary_currency": (org.settings or {}).get(
                        "primary_currency", request_currency
                    ),
                    "settings": org.settings,
                },
            }

        except Exception as e:
            await self.db.rollback()
            logger.error(f"Error completing onboarding: {e}", exc_info=True)
            raise

    async def calculate_temporary_bcr(self, request: TemporaryBCRRequest) -> dict[str, Any]:
        """
        Calculate BCR with temporary onboarding data (before saving)

        Args:
            request: TemporaryBCRRequest with temporary team and expense data

        Returns:
            Dictionary with calculated BCR and breakdown
        """
        request_currency = getattr(request.currency, "value", request.currency).upper()
        for member in request.team_members:
            member_currency = getattr(member.currency, "value", member.currency).upper()
            if member_currency != request_currency:
                raise ValueError(
                    f"Mixed currencies are not allowed in temporary BCR team members. "
                    f"Expected {request_currency}, got {member_currency}."
                )
        for expense in request.expenses:
            expense_currency = getattr(expense.currency, "value", expense.currency).upper()
            if expense_currency != request_currency:
                raise ValueError(
                    f"Mixed currencies are not allowed in temporary BCR expenses. "
                    f"Expected {request_currency}, got {expense_currency}."
                )
        for item in request.inventory_items or []:
            item_currency = str(item.get("currency") or "").upper()
            if item_currency and item_currency != request_currency:
                raise ValueError(
                    f"Mixed currencies are not allowed in temporary BCR inventory items. "
                    f"Expected {request_currency}, got {item_currency}."
                )

        # Calculate total salaries
        social_multiplier = Decimal("1")
        social_config = request.social_charges_config or {}
        if social_config.get("enable_social_charges"):
            total_percentage = Decimal("0")
            total_percentage += Decimal(str(social_config.get("health_percentage", 0) or 0))
            total_percentage += Decimal(str(social_config.get("pension_percentage", 0) or 0))
            total_percentage += Decimal(str(social_config.get("arl_percentage", 0) or 0))
            total_percentage += Decimal(str(social_config.get("parafiscales_percentage", 0) or 0))
            total_percentage += Decimal(str(social_config.get("prima_services_percentage", 0) or 0))
            total_percentage += Decimal(str(social_config.get("cesantias_percentage", 0) or 0))
            total_percentage += Decimal(str(social_config.get("int_cesantias_percentage", 0) or 0))
            total_percentage += Decimal(str(social_config.get("vacations_percentage", 0) or 0))
            if total_percentage == 0:
                total_percentage = Decimal(str(social_config.get("total_percentage", 0) or 0))
            if total_percentage > 0:
                social_multiplier = Decimal("1") + (total_percentage / Decimal("100"))

        total_salaries = sum(
            Decimal(str(member.salary_monthly_brute)) * social_multiplier
            for member in request.team_members
        )

        # Calculate total expenses
        total_expenses = sum(Decimal(str(expense.amount_monthly)) for expense in request.expenses)

        total_amortization = Decimal("0")
        for item in request.inventory_items or []:
            if not bool(item.get("amortizable")):
                continue
            purchase_price_raw = (
                item.get("purchase_price")
                if item.get("purchase_price") is not None
                else item.get("amount_monthly")
            )
            purchase_price = Decimal(str(purchase_price_raw or "0"))
            if purchase_price <= 0:
                continue
            salvage_value = Decimal(str(item.get("salvage_value") or "0"))
            if salvage_value < 0:
                salvage_value = Decimal("0")
            if salvage_value > purchase_price:
                salvage_value = purchase_price
            quantity = Decimal(str(item.get("quantity") or "1"))
            if quantity < 1:
                quantity = Decimal("1")
            category = str(item.get("category") or "").lower()
            useful_life_value = item.get("useful_life_months")
            if useful_life_value is None:
                useful_life_months = Decimal("24") if category == "software" else Decimal("36")
            else:
                useful_life_months = Decimal(str(useful_life_value or "0"))
            if useful_life_months <= 0:
                useful_life_months = Decimal("1")
            monthly = (purchase_price - salvage_value) / useful_life_months
            if monthly < 0:
                monthly = Decimal("0")
            total_amortization += monthly * quantity

        # Calculate total monthly costs
        total_monthly_costs = total_salaries + total_expenses + total_amortization

        # Keep parity with persisted model: monthly -> weekly int -> monthly (x 4.33)
        total_hours = Decimal("0")
        for member in request.team_members:
            weekly_hours = max(
                1, int(round(Decimal(str(member.billable_hours_per_month)) / Decimal("4.33")))
            )
            total_hours += Decimal(str(weekly_hours)) * Decimal("4.33")

        # Calculate BCR
        if total_hours > 0:
            bcr = total_monthly_costs / Decimal(str(total_hours))
        else:
            bcr = Decimal("0")

        return {
            "blended_cost_rate": str(bcr),
            "total_monthly_costs": str(total_monthly_costs),
            "total_fixed_overhead": str(total_expenses),
            "total_tools_costs": str(total_amortization),
            "total_salaries": str(total_salaries),
            "total_monthly_hours": float(total_hours),
            "team_members_count": len(request.team_members),
            "currency": request_currency,
            "note": "Values are calculated with temporary data and may differ after saving",
        }

    def _normalize_text(self, value: Any) -> str:
        if value is None:
            return ""
        return str(value).strip()

    def _normalize_upper(self, value: Any) -> str:
        return self._normalize_text(value).upper()

    def _issue(self, sheet: str, row: int, field: str, message: str) -> dict[str, Any]:
        return {"sheet": sheet, "row": row, "field": field, "message": message}

    def _parse_decimal(self, raw: Any) -> Decimal | None:
        if raw is None or str(raw).strip() == "":
            return None
        try:
            cleaned = str(raw).strip().replace(",", "")
            return Decimal(cleaned)
        except Exception:
            return None

    def _first_sheet_row(self, rows: list[dict[str, Any]]) -> dict[str, Any]:
        for row in rows:
            if any(self._normalize_text(value) for value in row.values()):
                return row
        return {}

    def _rows_from_openpyxl(self, worksheet) -> list[dict[str, Any]]:
        header: list[str] = []
        rows: list[dict[str, Any]] = []
        for idx, row in enumerate(worksheet.iter_rows(values_only=True), start=1):
            values = [self._normalize_text(cell) for cell in row]
            if idx == 1:
                header = values
                continue
            if not header:
                continue
            data = {header[i]: row[i] if i < len(row) else None for i in range(len(header))}
            if any(self._normalize_text(value) for value in data.values()):
                rows.append(data)
        return rows

    async def _build_payload_from_import_rows(
        self,
        organization_rows: list[dict[str, Any]],
        team_rows: list[dict[str, Any]],
        expense_rows: list[dict[str, Any]],
        inventory_rows: list[dict[str, Any]],
        source: str,
    ) -> dict[str, Any]:
        issues: list[dict[str, Any]] = []
        organization = self._first_sheet_row(organization_rows)

        organization_name = self._normalize_text(organization.get("organization_name"))
        country = self._normalize_upper(organization.get("country"))
        currency = self._normalize_upper(organization.get("currency"))
        profile_type = self._normalize_text(organization.get("profile_type")).lower() or "agency"
        organization_description = (
            self._normalize_text(organization.get("organization_description")) or None
        )

        if not country or country not in self.SUPPORTED_COUNTRY_CODES:
            issues.append(
                self._issue(
                    "Organization",
                    2,
                    "country",
                    "country must be one of COL, USA, ARG, MEX, PER, ESP",
                )
            )
        if not currency or currency not in self.SUPPORTED_CURRENCY_CODES:
            issues.append(
                self._issue(
                    "Organization",
                    2,
                    "currency",
                    "currency must be one of USD, COP, ARS, EUR, PEN, MXN",
                )
            )
        if profile_type not in {"freelance", "company", "agency"}:
            issues.append(
                self._issue(
                    "Organization",
                    2,
                    "profile_type",
                    "profile_type must be freelance, company, or agency",
                )
            )

        team_members: list[dict[str, Any]] = []
        for idx, row in enumerate(team_rows, start=2):
            name = self._normalize_text(row.get("name"))
            role = self._normalize_text(row.get("role"))
            salary = self._parse_decimal(row.get("salary_monthly_brute"))
            billable = row.get("billable_hours_per_month")
            member_currency = self._normalize_upper(row.get("currency")) or currency

            if not name:
                issues.append(self._issue("Team", idx, "name", "name is required"))
            if not role:
                issues.append(self._issue("Team", idx, "role", "role is required"))
            if salary is None or salary <= 0:
                issues.append(
                    self._issue(
                        "Team", idx, "salary_monthly_brute", "salary_monthly_brute must be > 0"
                    )
                )
            try:
                billable_int = int(str(billable).strip())
            except Exception:
                billable_int = 0
            if billable_int < 1 or billable_int > 200:
                issues.append(
                    self._issue(
                        "Team",
                        idx,
                        "billable_hours_per_month",
                        "billable_hours_per_month must be between 1 and 200",
                    )
                )
            if member_currency != currency:
                issues.append(
                    self._issue(
                        "Team",
                        idx,
                        "currency",
                        f"Mixed currencies are not allowed. Expected {currency}, got {member_currency}",
                    )
                )

            if (
                name
                and role
                and salary is not None
                and salary > 0
                and 1 <= billable_int <= 200
                and member_currency == currency
            ):
                team_members.append(
                    {
                        "name": name,
                        "role": role,
                        "salary_monthly_brute": str(salary),
                        "currency": member_currency,
                        "billable_hours_per_month": billable_int,
                    }
                )

        expenses: list[dict[str, Any]] = []
        for idx, row in enumerate(expense_rows, start=2):
            name = self._normalize_text(row.get("name"))
            category = self._normalize_text(row.get("category")).lower()
            amount = self._parse_decimal(row.get("amount_monthly"))
            item_currency = self._normalize_upper(row.get("currency")) or currency
            quantity_raw = row.get("quantity") or 1
            try:
                quantity = int(str(quantity_raw).strip())
            except Exception:
                quantity = 0

            if not name:
                issues.append(self._issue("Expenses", idx, "name", "name is required"))
            if category not in {"rent", "software", "services"}:
                issues.append(
                    self._issue(
                        "Expenses", idx, "category", "category must be rent, software, or services"
                    )
                )
            if amount is None or amount <= 0:
                issues.append(
                    self._issue("Expenses", idx, "amount_monthly", "amount_monthly must be > 0")
                )
            if quantity < 1:
                issues.append(self._issue("Expenses", idx, "quantity", "quantity must be >= 1"))
            if item_currency != currency:
                issues.append(
                    self._issue(
                        "Expenses",
                        idx,
                        "currency",
                        f"Mixed currencies are not allowed. Expected {currency}, got {item_currency}",
                    )
                )

            if (
                name
                and category in {"rent", "software", "services"}
                and amount is not None
                and amount > 0
                and quantity >= 1
                and item_currency == currency
            ):
                expenses.append(
                    {
                        "name": name,
                        "category": category,
                        "amount_monthly": str(amount),
                        "currency": item_currency,
                        "quantity": quantity,
                    }
                )

        inventory_items: list[dict[str, Any]] = []
        for idx, row in enumerate(inventory_rows, start=2):
            name = self._normalize_text(row.get("name"))
            category = self._normalize_text(row.get("category"))
            purchase_price = self._parse_decimal(row.get("purchase_price"))
            amount_monthly = self._parse_decimal(row.get("amount_monthly"))
            amount = purchase_price if purchase_price is not None else amount_monthly
            item_currency = self._normalize_upper(row.get("currency")) or currency
            amortizable_raw = self._normalize_upper(row.get("amortizable"))
            amortizable = amortizable_raw in {"TRUE", "1", "YES", "SI", "SÍ"}
            useful_life_raw = row.get("useful_life_months")
            salvage_value = self._parse_decimal(row.get("salvage_value")) or Decimal("0")
            purchase_date_raw = self._normalize_text(row.get("purchase_date")) or None
            depreciation_method = (
                self._normalize_text(row.get("depreciation_method")).lower() or "straight_line"
            )
            quantity_raw = row.get("quantity") or 1
            try:
                quantity = int(str(quantity_raw).strip())
            except Exception:
                quantity = 0
            try:
                useful_life_months = (
                    int(str(useful_life_raw).strip())
                    if useful_life_raw not in (None, "")
                    else (24 if category.lower() == "software" else 36)
                )
            except Exception:
                useful_life_months = 0

            if not name:
                issues.append(self._issue("Inventory", idx, "name", "name is required"))
            if amount is None or amount <= 0:
                issues.append(
                    self._issue(
                        "Inventory",
                        idx,
                        "purchase_price",
                        "purchase_price (or amount_monthly) must be > 0",
                    )
                )
            if quantity < 1:
                issues.append(self._issue("Inventory", idx, "quantity", "quantity must be >= 1"))
            if item_currency != currency:
                issues.append(
                    self._issue(
                        "Inventory",
                        idx,
                        "currency",
                        f"Mixed currencies are not allowed. Expected {currency}, got {item_currency}",
                    )
                )
            if amortizable and useful_life_months < 1:
                issues.append(
                    self._issue(
                        "Inventory",
                        idx,
                        "useful_life_months",
                        "useful_life_months must be >= 1 for amortizable items",
                    )
                )
            if salvage_value < 0:
                issues.append(
                    self._issue("Inventory", idx, "salvage_value", "salvage_value must be >= 0")
                )
            if amount is not None and salvage_value > amount:
                issues.append(
                    self._issue(
                        "Inventory",
                        idx,
                        "salvage_value",
                        "salvage_value cannot be greater than purchase_price",
                    )
                )
            if depreciation_method not in {"straight_line", "declining_balance"}:
                issues.append(
                    self._issue(
                        "Inventory",
                        idx,
                        "depreciation_method",
                        "depreciation_method must be straight_line or declining_balance",
                    )
                )

            if (
                name
                and amount is not None
                and amount > 0
                and quantity >= 1
                and item_currency == currency
                and (not amortizable or useful_life_months >= 1)
                and salvage_value >= 0
                and salvage_value <= amount
                and depreciation_method in {"straight_line", "declining_balance"}
            ):
                inventory_items.append(
                    {
                        "name": name,
                        "category": category or "Other",
                        "amount_monthly": str(amount_monthly or amount),
                        "purchase_price": str(amount),
                        "useful_life_months": useful_life_months,
                        "salvage_value": str(salvage_value),
                        "purchase_date": purchase_date_raw,
                        "depreciation_method": depreciation_method,
                        "currency": item_currency,
                        "quantity": quantity,
                        "amortizable": amortizable,
                    }
                )

        payload = None
        temporary_bcr = None
        if not issues:
            payload = {
                "organization_name": organization_name or None,
                "organization_description": organization_description,
                "country": country,
                "currency": currency,
                "profile_type": profile_type,
                "team_members": team_members,
                "expenses": expenses,
                "inventory_items": inventory_items,
            }
            if team_members:
                temp_request = TemporaryBCRRequest(
                    team_members=[OnboardingTeamMember(**member) for member in team_members],
                    expenses=[OnboardingExpense(**expense) for expense in expenses],
                    inventory_items=inventory_items,
                    currency=currency,
                )
                temporary_bcr = await self.calculate_temporary_bcr(temp_request)

        return {
            "success": len(issues) == 0,
            "source": source,
            "summary": {
                "team_members": len(team_members),
                "expenses": len(expenses),
                "inventory_items": len(inventory_items),
                "issues": len(issues),
            },
            "issues": issues,
            "payload": payload,
            "temporary_bcr": temporary_bcr,
        }

    async def preview_import_from_excel(self, file_content: bytes) -> dict[str, Any]:
        try:
            from openpyxl import load_workbook
        except Exception as exc:
            raise ValueError("openpyxl is required to import Excel files (.xlsx)") from exc

        workbook = load_workbook(filename=BytesIO(file_content), data_only=True)
        required_sheets = {"Organization", "Team", "Expenses", "Inventory"}
        existing_sheets = set(workbook.sheetnames)
        missing = required_sheets - existing_sheets
        if missing:
            raise ValueError(f"Missing required worksheets: {', '.join(sorted(missing))}")

        organization_rows = self._rows_from_openpyxl(workbook["Organization"])
        team_rows = self._rows_from_openpyxl(workbook["Team"])
        expense_rows = self._rows_from_openpyxl(workbook["Expenses"])
        inventory_rows = self._rows_from_openpyxl(workbook["Inventory"])

        return await self._build_payload_from_import_rows(
            organization_rows=organization_rows,
            team_rows=team_rows,
            expense_rows=expense_rows,
            inventory_rows=inventory_rows,
            source="excel",
        )

    async def preview_import_from_google_sheets(
        self,
        request: OnboardingImportSheetsRequest,
    ) -> dict[str, Any]:
        client = get_sheets_client()
        if not client:
            raise ValueError(
                "Could not authenticate with Google Sheets. Check service account configuration."
            )

        spreadsheet = client.open_by_key(request.spreadsheet_id)

        def get_rows(sheet_name: str) -> list[dict[str, Any]]:
            worksheet = spreadsheet.worksheet(sheet_name)
            return worksheet.get_all_records()

        organization_rows = get_rows(request.organization_sheet_name)
        team_rows = get_rows(request.team_sheet_name)
        expense_rows = get_rows(request.expenses_sheet_name)
        inventory_rows = get_rows(request.inventory_sheet_name)

        return await self._build_payload_from_import_rows(
            organization_rows=organization_rows,
            team_rows=team_rows,
            expense_rows=expense_rows,
            inventory_rows=inventory_rows,
            source="google_sheets",
        )

    def generate_excel_import_template(self) -> bytes:
        """
        Generate the official onboarding import template (.xlsx).
        """
        try:
            from openpyxl import Workbook
        except Exception as exc:
            raise ValueError("openpyxl is required to generate Excel templates (.xlsx)") from exc

        workbook = Workbook()
        workbook.remove(workbook.active)

        organization_sheet = workbook.create_sheet("Organization")
        organization_sheet.append(
            [
                "organization_name",
                "organization_description",
                "country",
                "currency",
                "profile_type",
            ]
        )
        organization_sheet.append(
            [
                "Mi Agencia",
                "Agencia de ejemplo",
                "COL",
                "COP",
                "agency",
            ]
        )

        team_sheet = workbook.create_sheet("Team")
        team_sheet.append(
            [
                "name",
                "role",
                "salary_monthly_brute",
                "currency",
                "billable_hours_per_month",
            ]
        )
        team_sheet.append(["Ana", "Developer", "5000000", "COP", "160"])

        expenses_sheet = workbook.create_sheet("Expenses")
        expenses_sheet.append(
            [
                "name",
                "category",
                "amount_monthly",
                "currency",
                "quantity",
            ]
        )
        expenses_sheet.append(["Arriendo oficina", "rent", "2200000", "COP", "1"])
        expenses_sheet.append(["Figma", "software", "120000", "COP", "1"])

        inventory_sheet = workbook.create_sheet("Inventory")
        inventory_sheet.append(
            [
                "name",
                "category",
                "purchase_price",
                "useful_life_months",
                "salvage_value",
                "purchase_date",
                "depreciation_method",
                "amount_monthly",
                "currency",
                "quantity",
                "amortizable",
            ]
        )
        inventory_sheet.append(
            [
                "Laptop",
                "Tools",
                "8000000",
                "36",
                "0",
                "2026-01-01",
                "straight_line",
                "",
                "COP",
                "1",
                "true",
            ]
        )
        inventory_sheet.append(
            [
                "Monitor",
                "Tools",
                "1200000",
                "36",
                "0",
                "2026-01-01",
                "straight_line",
                "",
                "COP",
                "2",
                "true",
            ]
        )
        inventory_sheet.append(
            ["Dominio web", "Other", "", "", "", "", "", "45000", "COP", "1", "false"]
        )

        notes_sheet = workbook.create_sheet("README")
        notes_sheet.append(["Campo", "Regla"])
        notes_sheet.append(["country", "Obligatorio. Valores: COL, USA, ARG, MEX, PER, ESP"])
        notes_sheet.append(
            [
                "currency",
                "Obligatorio y único para todo el archivo. Valores: USD, COP, ARS, EUR, PEN, MXN",
            ]
        )
        notes_sheet.append(["profile_type", "Obligatorio. Valores: freelance, company, agency"])
        notes_sheet.append(["expenses.category", "Obligatorio. Valores: rent, software, services"])
        notes_sheet.append(["amortizable", "Booleano: true/false, 1/0, yes/no"])
        notes_sheet.append(
            [
                "inventory.purchase_price",
                "Requerido para amortizables. No usar valor mensual para esos items",
            ]
        )
        notes_sheet.append(
            ["inventory.useful_life_months", "Requerido para amortizables. Debe ser >= 1"]
        )
        notes_sheet.append(
            [
                "inventory.salvage_value",
                "Opcional, por defecto 0. No puede ser mayor al purchase_price",
            ]
        )
        notes_sheet.append(["inventory.depreciation_method", "straight_line o declining_balance"])
        notes_sheet.append(["Importante", "No se permite mezcla de moneda entre hojas/filas"])

        output = BytesIO()
        workbook.save(output)
        output.seek(0)
        return output.getvalue()
